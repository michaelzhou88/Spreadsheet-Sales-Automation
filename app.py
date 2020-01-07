# Spreadsheet Automation Program
# Importing libraries
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Self defined function to reduce the prices of products in stock
def process_workbook(filename):
    # Loads the excel spreadsheet
    wb = xl.load_workbook(filename)
    # Returns the name of the sheet that needs to be automated
    sheet = wb['Sheet1']

    # Prints out the values of the initial prices
    print("Initial prices:")
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print("£" + str(cell.value))

    # Prints out values of the corrected prices into a new column
    print("Corrected prices:")
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = round((cell.value * 0.9), 2)
        correct_price_in_pounds = ("£" + str(corrected_price))
        print(correct_price_in_pounds)
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Select values to insert into bar chart
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    # Create bar chart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # Overwrites the original spreadsheet
    wb.save(filename)

# Initialising variable to specify the filename
filename = "transactions.xlsx"
# Invokes the process_workbook function
process_workbook(filename)